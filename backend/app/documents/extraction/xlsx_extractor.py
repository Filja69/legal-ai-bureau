"""XLSX extraction — best-effort, not load-bearing for completion (Phase 9.2
brief §3: "XLSX не должен блокировать completion"). Each sheet becomes one
`ExtractedTable`; cell values are stringified for the flat-text view.
"""
from __future__ import annotations

import zipfile
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.documents.extraction.base import ExtractedDocument, ExtractedSection, ExtractedTable, ExtractionError


class XlsxExtractor:
    async def extract(self, content: bytes) -> ExtractedDocument:
        try:
            workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        except (InvalidFileException, KeyError, ValueError, zipfile.BadZipFile) as exc:
            raise ExtractionError("CORRUPTED_FILE", "XLSX could not be parsed — the file may be corrupted") from exc

        tables: list[ExtractedTable] = []
        sections: list[ExtractedSection] = []
        for sheet in workbook.worksheets:
            rows: list[list[str]] = []
            for row in sheet.iter_rows():
                rows.append(["" if cell.value is None else str(cell.value) for cell in row])
            tables.append(ExtractedTable(page_number=None, rows=rows))
            sheet_text = "\n".join(" | ".join(row) for row in rows)
            sections.append(ExtractedSection(text=sheet_text, section_path=sheet.title))

        full_text = "\n\n".join(s.text for s in sections)
        return ExtractedDocument(
            text=full_text,
            sections=sections,
            tables=tables,
            metadata={"sheet_count": len(tables)},
            extractor="openpyxl",
        )
